# write excel
import json 
import os
import openpyxl
from  openpyxl import Workbook

# Import datetime class from datetime module
from datetime import datetime

from Config import *

cfg = Config()
record_root = cfg.record_root
 
# returns current date and time
now = datetime.now()
print(now)
print("using Record API")

def SaveJson(path, source_dict):
    jsonString = json.dumps(source_dict)
    jsonFile = open(path, "w")
    jsonFile.write(jsonString)
    jsonFile.close()

# def save_checkpoint(state, filename='checkpoint.pth.tar'):
    # # best_model = copy.deepcopy(model)
    # torch.save(state, filename)

def Write_Single_task_Training_Record(root = record_root, task = 2, epochs=100):

    # root = './record/'
    path = root +'/train_record.xlsx'

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)

    # read data from json
    json_path = root+'/task_{}_epoch_{}_history.json'.format(task, epochs)

    with open(json_path) as f:
        data = json.load(f)
        
    sheet_name= 'task_{}_training'.format(task)
    
    if sheet_name in wb.sheetnames:
        #load the sheet
        sh = wb[sheet_name]
    else:
        #create the sheet
        wb.create_sheet(sheet_name)
        sh = wb[sheet_name]

    
    # Write row name
    keys = ['loss', 'acc', 'loss1', 'loss2'] 
    for i in range(len(keys)):
     
        sh.cell(row = i+2 , column = 1, value = keys[i])
  
    # Write row of loss and acc
    record_len = len(data['loss'])
    
    for i in range(len(keys)):
    
        if keys[i] not in data.keys():
            continue
        else:
            for j in range(record_len):
        
                sh.cell(row = i+2 , column = j+2, value = data[keys[i]][j])



    wb.save(path)
    print("writing complete.")

def Write_Single_Task_TestAcc_Record(root = record_root, task_acc_list =[], train_task = 2):

    path = root+'/test_record.xlsx'.format(train_task)

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)


    if 'task_test' in wb.sheetnames:
        #load the sheet
        sh = wb['task_test']
    else:
        #create the sheet
        wb.create_sheet('task_test')
        sh = wb['task_test']
        
        for i in range(11):
            sh.cell(row = 1 , column = 2+i, value = i+1)
        
        for i in range(11):
            sh.cell(row = 2+i , column = 1, value = i+1)
       
    for i in range(len(task_acc_list)):
     
        sh.cell(row = i+2 , column = 2 + train_task - 1 , value = task_acc_list[i])


    wb.save(path)

def Write_Single_Task_TestPrec_Record(root = record_root, task_prec_list =[], train_task = 2):

    path = root+'/test_record.xlsx'.format(train_task)

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)


    if 'task_test_prec' in wb.sheetnames:
        #load the sheet
        sh = wb['task_test_prec']
    else:
        #create the sheet
        wb.create_sheet('task_test_prec')
        sh = wb['task_test_prec']
        
        for i in range(11):
            sh.cell(row = 1 , column = 2+i, value = i+1)
        
        for i in range(11):
            sh.cell(row = 2+i , column = 1, value = i+1)
       
    for i in range(len(task_prec_list)):
     
        sh.cell(row = i+2 , column = 2 + train_task - 1 , value = task_prec_list[i])


    wb.save(path)

def Write_Single_Task_CNNTestAcc_Record(root = record_root, task_acc_list =[], train_task = 2):

    path = root+'/test_record.xlsx'

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)


    if 'cnn_task_test' in wb.sheetnames:
        #load the sheet
        sh = wb['cnn_task_test']
    else:
        #create the sheet
        wb.create_sheet('cnn_task_test')
        sh = wb['cnn_task_test']
        
        for i in range(11):
            sh.cell(row = 1 , column = 2+i, value = i+1)
        
        for i in range(11):
            sh.cell(row = 2+i , column = 1, value = i+1)
       
    for i in range(len(task_acc_list)):
     
        sh.cell(row = i+2 , column = 2 + train_task - 1 , value = task_acc_list[i])


    wb.save(path)    

def Write_Single_Task_SSCNNTestAcc_Record(root = record_root, task_acc_list =[], train_task = 2):

    path = root+'/test_record.xlsx'

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)

    
    specific_sheetname = 'ss_cnn_task_test'
    
    if specific_sheetname in wb.sheetnames:
        #load the sheet
        sh = wb[specific_sheetname]
    else:
        #create the sheet
        wb.create_sheet(specific_sheetname)
        sh = wb[specific_sheetname]
        
        for i in range(11):
            sh.cell(row = 1 , column = 2+i, value = i+1)
        
        for i in range(11):
            sh.cell(row = 2+i , column = 1, value = i+1)
       
    for i in range(len(task_acc_list)):
     
        sh.cell(row = i+2 , column = 2 + train_task - 1 , value = task_acc_list[i])


    wb.save(path)        