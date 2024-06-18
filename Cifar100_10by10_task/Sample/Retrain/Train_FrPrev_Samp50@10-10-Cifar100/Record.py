# write excel
import json 
import os
import openpyxl
from  openpyxl import Workbook
from datetime import datetime

from Config import *

cfg = Config()
record_root = cfg.record_root


# returns current date and time
now = datetime.now()
print(now)
print("using Record API")

def SaveJson(path, source_dict):
    jsonString = json.dumps(source_dict)
    jsonFile = open(path, "w")
    jsonFile.write(jsonString)
    jsonFile.close()

def WriteExcel(data_dict, task = 2, item_names = ['max', 'min', 'mean', 'std'] ):

    #excel path
    root = cfg.record_root
    path = root +'task_sample_statistic.xlsx'
    # specify sheet name by task
    sheet_name = 'task{}_sample'.format(task)

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)
    

    if sheet_name in wb.sheetnames:
        #load the sheet
        sh = wb[sheet_name]
    else:
        #create the sheet
        wb.create_sheet(sheet_name)
        sh = wb[sheet_name]

    # read data from json
    # with open(json_path) as f:
        # data = json.load(f)
    
    #check if excel has been written    
    if sh.max_row != 0:
        start_row = sh.max_row + 2
    else:
        start_row = 0


    # Write row name in sheet
    keys = item_names
    for i in range(len(keys)):
     
        sh.cell(row = start_row + i , column = 1, value = keys[i])  
                    
    data = data_dict
                
    # Write row of loss and acc in sheet    
    for i in range(len(keys)):
    
        if keys[i] not in data.keys():
            continue
        else:
            record_len = len(data[keys[i]])
            for j in range(record_len):
                # print(data[keys[i]][j])
                sh.cell(row = start_row + i , column = j+2, value = data[keys[i]][j])
     
    #save excel
    wb.save(path)
    print("writing complete.")

def WriteStatExcel(data_dict):

    #excel path
    root = cfg.record_root
    path = root +'/recover_statistic.xlsx'
    # specify sheet name by task
    sheet_name = 'final_task_recover'

    if os.path.exists(root) ==False:
        os.mkdir(root)

    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)
    
    if sheet_name in wb.sheetnames:
        #load the sheet
        sh = wb[sheet_name]
    else:
        #create the sheet
        wb.create_sheet(sheet_name)
        sh = wb[sheet_name]

    # read data from json
    # with open(json_path) as f:
        # data = json.load(f)
    
    #check if excel has been written    
    if sh.max_row != 0:
        start_row = sh.max_row + 2
    else:
        start_row = 0


    # Write row name in sheet
    train_task = list(data_dict.keys())
    for i in range(len(train_task)):
        # print(train_task[i])
        sh.cell(row = start_row + 1 , column = i+1, value = train_task[i])  
                    
        data = data_dict[train_task[i]]
        
        #write horizontal attr
        for j, key in enumerate(list(data.keys())):
            
            sh.cell(row = start_row + 2 + j, column = i+1, value = data[key]["mean"])
 
    #save excel
    wb.save(path)
    print("writing complete.")

def Write_SingleTaskTrain_Record(task = 2, json_path = '', item_names = ['loss', 'acc', 'loss1', 'loss2'] ):
    
    print(task)

    #excel path
    root = cfg.record_root
    path = root + '/train_record.xlsx'
    
    # specify sheet name by task
    sheet_name = 'task_{}_training'.format(task)
    
    # open Excel and sheet
    wb, sh = OpenNewExcel(path, sheet_name)
   
    #check if excel has been written    
    if sh.max_row != 0:
        start_row = sh.max_row + 1
    else:
        start_row = 0
    
    #write time'stamp
    now = datetime.now()
    sh.cell(row = start_row  , column = 1, value = now)
    start_row += 1
     
    # read data from json
    with open(json_path) as f:
        data = json.load(f)  
    
    # print(json_path)
    
    # Write row name
    keys = item_names
    for i, k in enumerate(keys):
        Write_Line_AT_InExcel(title = k, values = data[k], wb_name = 'train_record.xlsx', sh_name = sheet_name,  mode = 'h', row = start_row+i, col = 1)
    
    print("writing complete.")
   
def Write_Single_task_Training_Record(task = 2, epochs = 100):

    root = cfg.record_root
    path = root + '/train_record.xlsx'

    # read data from json
    json_path = root+'/task_{}_epoch_{}_history.json'.format(task, epochs)

    with open(json_path) as f:
        data = json.load(f)        
    
    # Write row name
    keys = ['loss', 'acc', 'loss1', 'loss2']
    for k in keys:
        Write_Excel(title = k, values = data[k], wb_name = 'Exam_loss.xlsx', sh_name = 'exam1',  mode = 'h')
    
    print("writing complete.")

# following 5 function can be the same
def Write_Single_Task_Test_Result(task_acc_list =[], train_task = 2, total_task = 11, record_name = 'task_test'):
    
    root = cfg.record_root
    wb_name = 'test_record.xlsx'
    sheet_name = 'task_test'
    path = root + '/' + wb_name    
    
    wb, sh = OpenNewExcel(path, sheet_name)     
    
    titles =[]
    for i in range(total_task):
        titles.append(i+1)
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'h')
    Write_Line_AT_InExcel(title = None, values = task_acc_list, wb_name = wb_name , sh_name = sheet_name,  mode = 'v', row = 2, col = 2 + train_task-1)  
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'v')       
    
def Write_Single_Task_TestAcc_Record( task_acc_list =[], train_task = 2, total_task = 11):
    
    root = cfg.record_root

    wb_name = 'test_record.xlsx'
    sheet_name = 'task_test'
    path = root + '/' + wb_name    
    
    wb, sh = OpenNewExcel(path, sheet_name)     
    
    titles =[]
    for i in range(total_task):
        titles.append(i+1)
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'h')
    Write_Line_AT_InExcel(title = None, values = task_acc_list, wb_name = wb_name , sh_name = sheet_name,  mode = 'v', row = 2, col = 2 + train_task-1)  
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'v')

def Write_Single_Task_TestPrec_Record( task_prec_list =[], train_task = 2, total_task = 11):

    root = cfg.record_root
    wb_name = 'test_record.xlsx'
    sheet_name = 'task_prec'
    path = root + '/' + wb_name    
    
    titles =[]
    for i in range(total_task):
        titles.append(i+1)
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'h')
    Write_Line_AT_InExcel(title = None, values = task_prec_list, wb_name = wb_name , sh_name = sheet_name,  mode = 'v', row = 2, col = 2 + train_task-1)  
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'v')

def Write_Single_Task_CNNTestAcc_Record( task_acc_list =[], train_task = 2,  total_task = 11):
    
    root = cfg.record_root
    wb_name = 'test_record.xlsx'
    sheet_name = 'cnn_task_test'
    path = root + '/' + wb_name    
   
    titles =[]
    for i in range(total_task):
        titles.append(i+1)
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'h')
    Write_Line_AT_InExcel(title = None, values = task_acc_list, wb_name = wb_name , sh_name = sheet_name,  mode = 'v', row = 2, col = 2 + train_task-1)  
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'v')

def Write_Single_Task_SSCNNTestAcc_Record( task_acc_list =[], train_task = 2,  total_task = 11):
   
    root = cfg.record_root
    wb_name = 'test_record.xlsx'
    sheet_name = 'sscnn_task_test'
    path = root + '/' + wb_name    
   
    titles =[]
    for i in range(total_task):
        titles.append(i+1)
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'h')
    Write_Line_AT_InExcel(title = None, values = task_acc_list, wb_name = wb_name , sh_name = sheet_name,  mode = 'v', row = 2, col = 2 + train_task-1)  
    Write_Line_AT_InExcel(title = ' ', values = titles, wb_name = wb_name , sh_name = sheet_name,  mode = 'v')


def Write_Line_AT_InExcel(title = '', values = [], wb_name = '', sh_name = '',  mode = 'h', row = 1 , col = 1):

    path = record_root + '/' + wb_name
    sheet_name = sh_name
    wb, sh = OpenNewExcel(path, sheet_name)

    start_row = row
    start_col = col    
    
    # write horizontally: column change
    if mode == "h": 
        
        if title != None:
            sh.cell(row = start_row , column = start_col , value = title)
            start_col += 1
        
        for i, v in enumerate(values):
        
            sh.cell(row = start_row  , column = start_col + i  , value = v)

    # write vertically: row change
    if mode == "v": 
    
        if title != None:
            sh.cell(row = start_row , column = start_col , value = title)
            start_row += 1
        
        for i, v in enumerate(values):
        
            sh.cell(row = start_row + i  , column = start_col , value = v)   
    
    wb.save(path)              

def OpenNewExcel(path, sheet_name):
    
    if os.path.isfile(path):
        #load excel
        wb = openpyxl.load_workbook(path)
    else:
        #create excel
        wb = Workbook()
        wb.save(path)
        
    if sheet_name in wb.sheetnames:
        #load the sheet
        sh = wb[sheet_name]
    else:
        #create the sheet
        wb.create_sheet(sheet_name)
        sh = wb[sheet_name]
        
    return wb, sh    


if __name__ == "__main__":
    print(__name__)
    # for i in range(11):
        # Write_Excel(title = ' ', values = [str(i)], wb_name = 'test_record.xlsx' , sh_name = 'ss_cnn_task_test',  mode = 'h')
    
    # x = [1,2,4,5]
    # Write_Excel(title = ' ', values = x, wb_name = 'test_record.xlsx' , sh_name = 'ss_cnn_task_test',  mode = 'v')   